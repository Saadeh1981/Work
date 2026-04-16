#!/usr/bin/env python3
"""
Force-write PV Site - Metadata to fixed cells (first sheet): B4..B12
Fields/cells:
  B4  Plant Name
  B5  Country
  B6  AC Capacity (kWp)
  B7  DC Capacity (kW)
  B8  Active Power Export Power Limit (kW)
  B9  Latitude (WGS84)
  B10 Longitude (WGS84)
  B11 Elevation (m)
  B12 Commissioning/Comissioning date (yyyy-mm-dd)

This bypasses label lookup entirely.

Usage:
  python fill_site_metadata_to_cells.py `
    --input ".\\inputs\\your.pdf" `
    --template ".\\inputs\\PV metadata _TPM.xlsx" `
    --output ".\\out\\step2_site.xlsx" `
    --api-struct "http://127.0.0.1:8000/extract/asbuilt" `
    --api-ocr "http://127.0.0.1:8000/extract/read-test" `
    --first-pages 8 `
    --write-na

Requires: requests (or httpx), openpyxl, python-dateutil
"""
import argparse
import re
from pathlib import Path
from typing import Optional, Tuple, Any
import sys

try:
    import requests
except ImportError:
    import httpx as requests  # type: ignore

from openpyxl import load_workbook
from dateutil import parser as dateparser

DEFAULT_STRUCT = "http://127.0.0.1:8000/extract/asbuilt"
DEFAULT_OCR    = "http://127.0.0.1:8000/extract/read-test"

def fetch_struct(api: str, pdf: Path, first_pages: int):
    files = {"file": (pdf.name, open(pdf, "rb"), "application/pdf")}
    params = {"first_pages": first_pages}
    try:
        r = requests.post(api, files=files, params=params, timeout=240)
    finally:
        try: files["file"][1].close()
        except Exception: pass
    try:
        return r.json() if r.status_code == 200 else {}
    except Exception:
        return {}

def fetch_text(api: str, pdf: Path, first_pages: int) -> str:
    files = {"file": (pdf.name, open(pdf, "rb"), "application/pdf")}
    params = {"first_pages": first_pages}
    try:
        r = requests.post(api, files=files, params=params, timeout=240)
    finally:
        try: files["file"][1].close()
        except Exception: pass
    try:
        if r.status_code != 200: return ""
        data = r.json()
    except Exception:
        try: return r.text
        except Exception: return ""
    if isinstance(data, dict):
        for k in ["text","full_text","content","ocr","raw_text"]:
            if k in data and isinstance(data[k], str):
                return data[k]
        parts = []
        for v in data.values():
            if isinstance(v, str): parts.append(v)
            elif isinstance(v, list): parts += [x for x in v if isinstance(x, str)]
        return "\n".join(parts)
    return str(data)

def norm(x: Any) -> str:
    return "" if x is None else str(x).strip()

def mw_to_kw(x: Optional[float]) -> Optional[float]:
    try:
        return None if x is None else float(x)*1000.0
    except Exception:
        return None

def parse_ac_kw(U: str) -> Optional[float]:
    pats = [
        r"\bAC\s*(?:CAPACITY|NAMEPLATE|SYSTEM|RATING)[^0-9]{0,12}([0-9]+(?:\.[0-9]+)?)\s*(MW|KW)\b",
        r"([0-9]+(?:\.[0-9]+)?)\s*(KWAC|KW AC|KVA)\b",
    ]
    for p in pats:
        m = re.search(p, U, re.I)
        if m:
            val = float(m.group(1))
            unit = (m.group(2) or "KW").upper().replace(" ","")
            return val*1000.0 if unit in ("MW","MVA") else val
    return None

def parse_dc_kw(U: str) -> Optional[float]:
    pats = [
        r"\bDC\s*(?:CAPACITY|NAMEPLATE|SYSTEM|RATING|STC)[^0-9]{0,12}([0-9]+(?:\.[0-9]+)?)\s*(MW|KW|KWP|KWDC)\b",
        r"([0-9]+(?:\.[0-9]+)?)\s*(KWDC|KW DC|KWP)\b",
        r"TOTAL\s*DC\s*(?:SYSTEM)?\s*RATING[^0-9]{0,12}([0-9]+(?:\.[0-9]+)?)\s*(KWDC|KWP|KW)\b",
    ]
    for p in pats:
        m = re.search(p, U, re.I)
        if m:
            val = float(m.group(1))
            unit = (m.group(2) or "KW").upper().replace(" ","")
            return val*1000.0 if unit == "MW" else val
    return None

def parse_export_limit(U: str) -> Optional[float]:
    pats = [
        r"(?:EXPORT|ACTIVE\s*POWER|POWER\s*EXPORT).{0,40}LIMIT[^0-9]{0,10}([0-9]+(?:\.[0-9]+)?)\s*(MW|KW)?",
        r"LIMIT\s*(?:OF|:)?\s*([0-9]+(?:\.[0-9]+)?)\s*(MW|KW)?\b",
    ]
    for p in pats:
        m = re.search(p, U, re.I)
        if m:
            val = float(m.group(1)); unit = (m.group(2) or "KW").upper()
            return val*1000.0 if unit=="MW" else val
    return None

def parse_lat_lon(text: str) -> Tuple[Optional[float], Optional[float]]:
    lat = lon = None
    for line in text.splitlines():
        U = line.upper()
        mlat = re.search(r"LAT(?:ITUDE)?\s*[:=]?\s*([\-+]?\d{1,2}\.\d+)", U)
        mlon = re.search(r"LON(?:GITUDE)?\s*[:=]?\s*([\-+]?\d{1,3}\.\d+)", U)
        if mlat: lat = float(mlat.group(1))
        if mlon: lon = float(mlon.group(1))
        if lat is not None and lon is not None: return lat, lon
    for line in text.splitlines():
        nums = re.findall(r"[\-+]?\d{1,3}\.\d{3,}", line)
        if len(nums) >= 2:
            a, b = float(nums[0]), float(nums[1])
            if abs(a) <= 90 and abs(b) <= 180: return a, b
    return None, None

def parse_elevation(U: str) -> Optional[float]:
    m = re.search(r"ELEV(?:ATION)?\s*[:=]?\s*([0-9]+(?:\.[0-9]+)?)\s*(M|METERS|FT|FEET)?", U)
    if m:
        val = float(m.group(1)); unit = (m.group(2) or "M").upper()
        return val*0.3048 if unit in ("FT","FEET") else val
    return None

def parse_country(U: str) -> Optional[str]:
    if "PUERTO RICO" in U or " PR" in U or ",PR" in U: return "Puerto Rico"
    if "UNITED STATES" in U or " USA" in U: return "United States"
    return None

def parse_commissioning_date(text: str) -> Optional[str]:
    lines = text.splitlines()
    CUE = re.compile(r"(commission\w+|commercial\s*operation\s*date|cod)", re.I)
    for i, line in enumerate(lines):
        if CUE.search(line):
            near = " ".join(lines[max(0, i-1):min(len(lines), i+2)])
            pm = re.search(r"(\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|[A-Za-z]{3,9}\s+\d{1,2},\s*\d{4})", near)
            if pm:
                try:
                    dt = dateparser.parse(pm.group(1), dayfirst=False, yearfirst=False)
                    return dt.strftime("%Y-%m-%d")
                except Exception:
                    pass
    return None

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--template", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--api-struct", default=DEFAULT_STRUCT)
    ap.add_argument("--api-ocr", default=DEFAULT_OCR)
    ap.add_argument("--first-pages", type=int, default=8)
    ap.add_argument("--write-na", action="store_true")
    args = ap.parse_args()

    in_path = Path(args.input).resolve()
    pdf = in_path if in_path.is_file() else sorted(in_path.glob("**/*.pdf"))[0]

    struct = fetch_struct(args.api_struct, pdf, args.first_pages)
    text = fetch_text(args.api_ocr, pdf, args.first_pages)
    U = text.upper()

    plant_name = struct.get("plant_name") or ""
    ac_kwp = struct.get("ac_capacity_mw"); ac_kwp = mw_to_kw(ac_kwp) if ac_kwp is not None else parse_ac_kw(U)
    dc_kw  = struct.get("dc_capacity_mw"); dc_kw  = mw_to_kw(dc_kw)  if dc_kw  is not None else parse_dc_kw(U)
    export_kw = parse_export_limit(U)
    lat, lon  = parse_lat_lon(text)
    elev_m    = parse_elevation(U)
    country   = parse_country(U)
    cod       = parse_commissioning_date(text)

    # Write
    from openpyxl import load_workbook
    wb = load_workbook(Path(args.template).resolve())
    ws = wb.worksheets[0]
    def put(cell, val):
        v = "N/A" if (args.write_na and (val is None or val=="" )) else val
        ws[cell] = v
        print(f"WRITE {cell} = {v!r}")

    put("B4",  plant_name or "")
    put("B5",  country)
    put("B6",  None if ac_kwp is None else round(ac_kwp, 1))
    put("B7",  None if dc_kw  is None else round(dc_kw, 1))
    put("B8",  export_kw)
    put("B9",  lat)
    put("B10", lon)
    put("B11", elev_m)
    put("B12", cod)

    outp = Path(args.output).resolve()
    outp.parent.mkdir(parents=True, exist_ok=True)
    wb.save(outp)
    print(f"Saved -> {outp}")

    # Verify by reading back
    wb2 = load_workbook(outp, data_only=True)
    ws2 = wb2.worksheets[0]
    for ref in ["B4","B5","B6","B7","B8","B9","B10","B11","B12"]:
        print(f"READBACK {ref} -> {ws2[ref].value!r}")

if __name__ == "__main__":
    main()
