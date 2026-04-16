#!/usr/bin/env python3
"""
Scan ALL pages of a PDF, extract PV Site - Metadata, and write to the first sheet:
  B4  Plant Name
  B5  Country
  B6  AC Capacity (kWp)
  B7  DC Capacity (kW)
  B8  Active Power Export Power Limit (kW)
  B9  Latitude (WGS84)
  B10 Longitude (WGS84)
  B11 Elevation (m)
  B12 Commissioning date (yyyy-mm-dd)
"""
import argparse, io, re, sys
from pathlib import Path
from typing import Optional, Tuple, Any, List

# deps
try:
    import fitz  # PyMuPDF
except ImportError:
    print("ERROR: PyMuPDF not installed. Run:  python -m pip install PyMuPDF", file=sys.stderr)
    raise
try:
    import requests
except ImportError:
    import httpx as requests  # type: ignore
from openpyxl import load_workbook
from dateutil import parser as dateparser

DEFAULT_OCR = "http://127.0.0.1:8000/extract/read-test"

def post_pdf_bytes(api: str, name: str, data: bytes, timeout: int = 300) -> str:
    files = {"file": (name, io.BytesIO(data), "application/pdf")}
    r = requests.post(api, files=files, timeout=timeout)
    try:
        if r.status_code != 200:
            return ""
        obj = r.json()
    except Exception:
        return r.text if hasattr(r, "text") else ""
    if isinstance(obj, dict):
        for k in ["text","full_text","content","ocr","raw_text"]:
            if k in obj and isinstance(obj[k], str):
                return obj[k]
        parts: List[str] = []
        for v in obj.values():
            if isinstance(v, str): parts.append(v)
            elif isinstance(v, list): parts += [x for x in v if isinstance(x, str)]
        return "\n".join(parts)
    return str(obj)

def chunk_pdf_to_bytes(pdf_path: Path, chunk_pages: int):
    doc = fitz.open(pdf_path)
    n = doc.page_count
    for start in range(0, n, chunk_pages):
        end = min(n-1, start + chunk_pages - 1)
        out = fitz.open()
        out.insert_pdf(doc, from_page=start, to_page=end)
        yield (start+1, end+1, out.tobytes())
    doc.close()

def pick_lines(text: str, *needles: str, window: int = 0) -> List[str]:
    lines = text.splitlines()
    outs: List[str] = []
    pats = [re.compile(rf"\b{re.escape(n)}\b", re.I) for n in needles]
    for i, line in enumerate(lines):
        if any(p.search(line) for p in pats):
            if window == 0:
                outs.append(line)
            else:
                a = max(0, i-window); b = min(len(lines), i+window+1)
                outs.append(" ".join(lines[a:b]))
    return outs

def _num(s: str) -> float:
    return float(s.replace(",", ".").replace(" ", ""))

def parse_country(U: str) -> Optional[str]:
    if "PUERTO RICO" in U or " PR" in U or ",PR" in U: return "Puerto Rico"
    if "UNITED STATES" in U or " USA" in U: return "United States"
    if "MEXICO" in U: return "Mexico"
    if "CANADA" in U: return "Canada"
    return None

def parse_commissioning_date(text: str) -> Optional[str]:
    lines = text.splitlines()
    CUE = re.compile(r"(commission\w+|commercial\s*operation\s*date|cod)\b", re.I)
    for i, line in enumerate(lines):
        if CUE.search(line):
            near = " ".join(lines[max(0, i-1):min(len(lines), i+2)])
            pm = re.search(r"(\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|[A-Za-z]{3,9}\s+\d{1,2},\s*\d{4})", near)
            if pm:
                try:
                    dt = dateparser.parse(pm.group(1), dayfirst=False, yearfirst=False)
                    return dt.strftime("%Y-%m-%d")
                except Exception:
                    pass
    return None

def parse_lat_lon(text: str) -> Tuple[Optional[float], Optional[float]]:
    for line in text.splitlines():
        mlat = re.search(r"LAT(?:ITUDE)?\s*[:=]?\s*([\-+]?\d{1,2}[.,]\d+)", line, re.I)
        mlon = re.search(r"LON(?:GITUDE)?\s*[:=]?\s*([\-+]?\d{1,3}[.,]\d+)", line, re.I)
        if mlat and mlon:
            try:
                a = float(mlat.group(1).replace(",", ".")); b = float(mlon.group(1).replace(",", "."))
                return a, b
            except Exception:
                pass
    for line in text.splitlines():
        nums = re.findall(r"[\-+]?\d{1,3}[.,]\d{3,}", line)
        if len(nums) >= 2:
            try:
                a = float(nums[0].replace(",", ".")); b = float(nums[1].replace(",", "."))
                if abs(a) <= 90 and abs(b) <= 180: return a, b
            except Exception:
                pass
    return None, None

def parse_elevation(U: str) -> Optional[float]:
    m = re.search(r"ELEV(?:ATION)?\s*[:=]?\s*([0-9]+(?:[.,][0-9]+)?)\s*(M|METERS|FT|FEET)?\b", U, re.I)
    if m:
        val = _num(m.group(1)); unit = (m.group(2) or "M").upper()
        return val*0.3048 if unit in ("FT","FEET") else val
    return None

def parse_dc_kw(U: str) -> Optional[float]:
    pats = [
        r"TOTAL\s*DC\s*(?:SYSTEM)?\s*(?:RATING|POWER|CAPACITY|STC)[^0-9]{0,40}([0-9]+(?:[.,][0-9]+)?)\s*(MWDC|KWDC|KWP|MW|KW)\b",
        r"(?:PV\s*)?(?:SYSTEM|PLANT)\s*SIZE\s*\(DC\)\s*[:=]?\s*([0-9]+(?:[.,][0-9]+)?)\s*(MW|KW|KWP|KWDC)\b",
        r"\bDC\s*(?:CAPACITY|NAMEPLATE|SYSTEM|RATING|STC)[^0-9]{0,30}([0-9]+(?:[.,][0-9]+)?)\s*(MW|KW|MWDC|KWDC|KWP)\b",
        r"\b([0-9]+(?:[.,][0-9]+)?)\s*(MWDC|KWDC|KWP|MW-DC|KW-DC|KW DC|MW DC)\b",
    ]
    vals: List[float] = []
    for p in pats:
        for m in re.finditer(p, U, re.I):
            val = _num(m.group(1)); unit = m.group(2).upper().replace(" ", "").replace("-", "")
            kw = val*1000.0 if unit.startswith("MW") else val
            vals.append(kw)
    return max(vals) if vals else None

def parse_ac_kw(U: str) -> Optional[float]:
    pats = [
        r"(?:PV\s*)?(?:SYSTEM|PLANT)\s*SIZE\s*\(AC\)\s*[:=]?\s*([0-9]+(?:[.,][0-9]+)?)\s*(MW|KW|MWAC|KWAC|KVA)\b",
        r"\bAC\s*(?:CAPACITY|NAMEPLATE|SYSTEM|RATING)[^0-9]{0,30}([0-9]+(?:[.,][0-9]+)?)\s*(MWAC|KWAC|MW|KW|KVA)\b",
        r"\b([0-9]+(?:[.,][0-9]+)?)\s*(MWAC|KWAC|MW-AC|KW-AC|MW AC|KW AC|KVA)\b",
    ]
    vals: List[float] = []
    for p in pats:
        for m in re.finditer(p, U, re.I):
            val = _num(m.group(1)); unit = m.group(2).upper().replace(" ", "").replace("-", "")
            vals.append(val*1000.0 if unit.startswith("MW") else val)
    return max(vals) if vals else None

def parse_export_limit(U: str) -> Optional[float]:
    pats = [
        r"(?:EXPORT|INTERCONNECTION|ACTIVE\s*POWER|POWER\s*EXPORT).{0,80}LIMIT[^0-9]{0,20}([0-9]+(?:[.,][0-9]+)?)\s*(MW|KW)?\b",
        r"LIMIT\s*(?:OF|:)?\s*([0-9]+(?:[.,][0-9]+)?)\s*(MW|KW)?\b",
    ]
    for p in pats:
        m = re.search(p, U, re.I)
        if m:
            val = _num(m.group(1)); unit = (m.group(2) or "KW").upper()
            return val*1000.0 if unit == "MW" else val
    return None

def best_guess_plant_name(text: str, filename: str) -> str:
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    for i, line in enumerate(lines[:300]):
        U = line.upper()
        if "SOLAR" in U and "ELECTRIC" in U and "SYSTEM" in U:
            tail = " ".join(lines[i:i+3])
            m = re.search(r"SOLAR\s+ELECTRIC\s+SYSTEM\s+([^\n\r,]+)", tail, re.I)
            if m:
                cand = m.group(1).strip(" -,:")
                if 3 <= len(cand) <= 80: return cand.title()
    for line in lines[:200]:
        m = re.search(r"\b(PROJECT|PLANT)\s*[:\-]\s*([A-Za-z0-9 \-_'().,/]+)", line, re.I)
        if m:
            cand = m.group(2).strip()
            if 3 <= len(cand) <= 80: return cand
    name = Path(filename).stem.replace("_"," ").replace("-"," ")
    return name[:64]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--template", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--api-ocr", default=DEFAULT_OCR)
    ap.add_argument("--chunk-pages", type=int, default=10)
    ap.add_argument("--write-na", action="store_true")
    ap.add_argument("--dump-text", default=r".\out\merged_ocr.txt")
    args = ap.parse_args()

    pdf_path = Path(args.input).resolve()
    tpl = Path(args.template).resolve()
    outp = Path(args.output).resolve()

    merged: List[str] = []
    for i, (a,b,data) in enumerate(chunk_pdf_to_bytes(pdf_path, args.chunk_pages), 1):
        print(f"OCR chunk {i}: pages {a}-{b}")
        t = post_pdf_bytes(args.api_ocr, f"chunk_{a}-{b}.pdf", data)
        if t: merged.append(t)
    full_text = "\n".join(merged)
    Path(args.dump_text).parent.mkdir(parents=True, exist_ok=True)
    Path(args.dump_text).write_text(full_text, encoding="utf-8")
    print(f"Saved merged OCR to: {Path(args.dump_text).resolve()}")

    U = full_text.upper()
    plant_name = best_guess_plant_name(full_text, pdf_path.name)
    country   = parse_country(U)
    ac_kwp    = parse_ac_kw(U)
    dc_kw     = parse_dc_kw(U)
    export_kw = parse_export_limit(U)
    lat, lon  = parse_lat_lon(full_text)
    elev_m    = parse_elevation(U)
    cod       = parse_commissioning_date(full_text)

    previews = {
        "AC lines": pick_lines(full_text, "AC", "KWAC", "MWAC", "KVA", "AC CAPACITY", "SYSTEM SIZE (AC)"),
        "DC lines": pick_lines(full_text, "DC", "KWDC", "KWP", "TOTAL DC", "STC", "RATING", "SYSTEM SIZE (DC)"),
        "EXPORT lines": pick_lines(full_text, "EXPORT", "INTERCONNECTION", "LIMIT"),
    }
    for tag, hits in previews.items():
        if hits:
            print(f"\n[{tag}]")
            for h in hits[:8]:
                print("  ", h)

    print("\n=== DETECTED (all pages) ===")
    print(f"Plant Name: {plant_name}")
    print(f"Country: {country}")
    print(f"AC (kWp): {ac_kwp}")
    print(f"DC (kW): {dc_kw}")
    print(f"Export Limit (kW): {export_kw}")
    print(f"Lat/Lon: {lat}, {lon}")
    print(f"Elevation (m): {elev_m}")
    print(f"Commissioning: {cod}")
    print("============================")

    wb = load_workbook(tpl)
    ws = wb.worksheets[0]
    def put(cell, val):
        v = "N/A" if (args.write_na and (val is None or val == "")) else val
        ws[cell] = v
        print(f"WRITE {cell} = {v!r}")

    def r(x): return round(x,1) if isinstance(x,(int,float)) else x
    put("B4", plant_name or ""); put("B5", country); put("B6", r(ac_kwp)); put("B7", r(dc_kw))
    put("B8", r(export_kw)); put("B9", lat); put("B10", lon); put("B11", r(elev_m)); put("B12", cod)

    outp.parent.mkdir(parents=True, exist_ok=True)
    wb.save(outp)
    print(f"Saved -> {outp}")

if __name__ == "__main__":
    main()
