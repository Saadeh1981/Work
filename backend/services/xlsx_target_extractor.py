from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.services.evidence_builder import build_evidence
from backend.services.confidence_builder import score_text_block

def extract_xlsx_sheet(path: str, sheet_name: str, max_rows: int = 200) -> dict[str, Any]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "sheet_name": sheet_name,
            "headers": [],
            "rows": [],
            "row_count": 0,
            "row_evidence": [],
        }

    if sheet_name not in wb.sheetnames:
        return {
            "status": "error",
            "error": f"sheet not found: {sheet_name}",
            "sheet_name": sheet_name,
            "headers": [],
            "rows": [],
            "row_count": 0,
            "row_evidence": [],
        }

    ws = wb[sheet_name]
    file_name = Path(path).name

    raw_rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        values = [clean_cell(v) for v in row]
        if any(v != "" for v in values):
            raw_rows.append(values)
        if len(raw_rows) >= max_rows:
            break

    if not raw_rows:
        return {
            "status": "ok",
            "sheet_name": sheet_name,
            "headers": [],
            "rows": [],
            "row_count": 0,
            "row_evidence": [],
        }

    header_idx, headers = pick_header_row(raw_rows)

    data_rows: list[dict[str, str]] = []
    row_evidence: list[dict[str, Any]] = []

    for row_offset, row in enumerate(raw_rows[header_idx + 1:], start=header_idx + 1):
        row_dict: dict[str, str] = {}

        for i, header in enumerate(headers):
            key = header if header else f"column_{i+1}"
            value = row[i] if i < len(row) else ""
            row_dict[key] = value

        data_rows.append(row_dict)

        row_text_parts = []
        for key, value in row_dict.items():
            if value != "":
                row_text_parts.append(f"{key}: {value}")

        row_text = " | ".join(row_text_parts)

        row_evidence.append(
            build_evidence(
                file_name=file_name,
                file_type="xlsx",
                sheet_name=sheet_name,
                snippet=row_text,
                method="xlsx_row",
                block_index=row_offset,
                page=None,
                confidence=score_text_block(
                    text=row_text,
                    method="xlsx_row",
                    file_type="xlsx",
                ),
            )
        )

    return {
        "status": "ok",
        "sheet_name": sheet_name,
        "headers": headers,
        "rows": data_rows,
        "row_count": len(data_rows),
        "row_evidence": row_evidence,
    }


def pick_header_row(rows: list[list[str]]) -> tuple[int, list[str]]:
    best_idx = 0
    best_row = rows[0]
    best_score = -1

    for idx, row in enumerate(rows[:10]):
        score = score_header_row(row)
        if score > best_score:
            best_score = score
            best_row = row
            best_idx = idx

    return best_idx, normalize_headers(best_row)


def score_header_row(row: list[str]) -> int:
    non_empty = sum(1 for v in row if v)
    alpha_cells = sum(1 for v in row if any(ch.isalpha() for ch in v))
    unique_vals = len(set(v for v in row if v))
    return non_empty + alpha_cells + unique_vals


def normalize_headers(row: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []

    for i, value in enumerate(row):
        name = value.strip() if value else f"column_{i+1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        result.append(name)

    return result


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()