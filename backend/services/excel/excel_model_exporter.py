from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from backend.services.excel.device_sheet_schema import get_sheet_columns
from backend.services.excel.excel_row_builders import (
    build_device_rows_by_sheet,
    build_info1_rows,
    build_info2_rows,
    build_missing_questions_rows,
    get_present_device_sheets,
)

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
REVIEW_FILL = PatternFill(fill_type="solid", fgColor="FFF3CD")


class ExcelModelExporter:
    def __init__(self) -> None:
        pass

    def export(self, output: Dict[str, Any], output_path: str | Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        default_ws = wb.active
        wb.remove(default_ws)

        self._write_info1_sheet(wb, output)
        self._write_info2_sheet(wb, output)
        self._write_missing_questions_sheet(wb, output)

        rows_by_sheet = build_device_rows_by_sheet(output)
        present_sheets = get_present_device_sheets(output)

        self._write_review_summary(wb, rows_by_sheet)

        for sheet_name in present_sheets:
            self._write_device_sheet(
                wb=wb,
                sheet_name=sheet_name,
                rows=rows_by_sheet.get(sheet_name, []),
            )

        wb.save(output_path)
        return output_path

    def _write_info1_sheet(self, wb: Workbook, output: Dict[str, Any]) -> None:
        rows = build_info1_rows(output)
        ws = wb.create_sheet(title="Info 1")
        headers = ["Section", "Key", "Value"]
        self._write_rows(ws, headers, rows)

    def _write_info2_sheet(self, wb: Workbook, output: Dict[str, Any]) -> None:
        rows = build_info2_rows(output)
        ws = wb.create_sheet(title="Info 2")
        headers = ["Device Type Name", "Device Model Name", "Device Manufacturer"]
        self._write_rows(ws, headers, rows)

    def _write_missing_questions_sheet(self, wb: Workbook, output: Dict[str, Any]) -> None:
        rows = build_missing_questions_rows(output)
        ws = wb.create_sheet(title="Missing Questions")

        headers = [
            "Plant ID",
            "Scope",
            "Node ID",
            "Field",
            "Reason",
            "Priority",
            "Question for User",
            "User Answer",
            "Status",
        ]

        self._write_rows(ws, headers, rows)

        self._add_missing_question_status_validation(
            ws=ws,
            headers=headers,
            max_row=len(rows) + 1,
        )

    def _write_review_summary(
        self,
        wb: Workbook,
        rows_by_sheet: Dict[str, List[Dict[str, Any]]],
    ) -> None:
        summary_rows: List[Dict[str, Any]] = []

        for sheet_name, rows in rows_by_sheet.items():
            total_rows = len(rows)
            review_rows = sum(1 for r in rows if r.get("Review Required"))
            review_pct = round((review_rows / total_rows) * 100, 1) if total_rows else 0.0

            summary_rows.append(
                {
                    "Sheet": sheet_name,
                    "Total Rows": total_rows,
                    "Needs Review": review_rows,
                    "Review %": review_pct,
                }
            )

        summary_rows = sorted(summary_rows, key=lambda r: r["Sheet"])

        ws = wb.create_sheet(title="Review Summary")
        headers = ["Sheet", "Total Rows", "Needs Review", "Review %"]
        self._write_rows(ws, headers, summary_rows)

    def _write_device_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        rows: List[Dict[str, Any]],
    ) -> None:
        ws = wb.create_sheet(title=sheet_name)
        headers = get_sheet_columns(sheet_name)

        rows = sorted(
            rows,
            key=lambda r: (
                str(r.get("Parent", "")),
                str(r.get("Name", "")),
            ),
        )

        self._write_rows(ws, headers, rows)

        if "Modification status" in headers:
            self._add_status_validation(
                ws,
                headers.index("Modification status") + 1,
                len(rows) + 1,
            )

        self._highlight_review_rows(ws, headers)

    def _write_rows(self, ws, headers: List[str], rows: List[Dict[str, Any]]) -> None:
        ws.append(headers)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = WRAP_ALIGNMENT

        for row_data in rows:
            ws.append([row_data.get(header, "") for header in headers])

        self._format_sheet(ws, headers)

    def _format_sheet(self, ws, headers: List[str]) -> None:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wrap_columns = {
            "Review Reasons",
            "Evidence Summary",
            "Question for User",
            "User Answer",
        }

        for col_idx, header in enumerate(headers, start=1):
            max_len = len(header)

            for row_idx in range(2, min(ws.max_row, 300) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is not None:
                    max_len = max(max_len, len(str(value)))

                if header in wrap_columns:
                    cell.alignment = WRAP_ALIGNMENT
                else:
                    cell.alignment = Alignment(vertical="top")

            ws.cell(row=1, column=col_idx).alignment = WRAP_ALIGNMENT

            width = min(max(max_len + 2, 14), 40)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    def _highlight_review_rows(self, ws, headers: List[str]) -> None:
        if "Review Required" not in headers:
            return

        review_col_idx = headers.index("Review Required") + 1

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            value = row[review_col_idx - 1].value
            if value is True or str(value).strip().lower() in {"true", "yes", "1"}:
                for cell in row:
                    cell.fill = REVIEW_FILL

    def _add_status_validation(self, ws, column_index: int, max_row: int) -> None:
        dv = DataValidation(
            type="list",
            formula1='"New,Unchanged,Needs Review,Accepted,Corrected,Rejected,Deferred"',
            allow_blank=True,
        )
        ws.add_data_validation(dv)

        if max_row >= 2:
            col_letter = ws.cell(row=2, column=column_index).column_letter
            dv.add(f"{col_letter}2:{col_letter}{max_row}")


    def _add_missing_question_status_validation(
        self,
        ws,
        headers: List[str],
        max_row: int,
    ) -> None:
        if "Status" not in headers:
            return

        column_index = headers.index("Status") + 1

        dv = DataValidation(
            type="list",
            formula1='"Open,Answered,Deferred"',
            allow_blank=True,
        )
        ws.add_data_validation(dv)

        if max_row >= 2:
            col_letter = ws.cell(row=2, column=column_index).column_letter
            dv.add(f"{col_letter}2:{col_letter}{max_row}")