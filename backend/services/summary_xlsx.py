from __future__ import annotations

from collections import Counter
from typing import Iterable

from openpyxl import load_workbook

from backend.schemas.summary import FileSection, FileSummary
from backend.services.preflight_service import DocProfile


CATEGORY_RULES = {
    "plant_metadata": {
        "sheet_keywords": {
            "metadata", "plant", "site", "project", "facility", "overview", "general"
        },
        "header_keywords": {
            "plant", "site", "project", "facility", "location", "address",
            "latitude", "longitude", "utility", "interconnection", "cod",
            "commissioning", "owner", "customer", "portfolio", "technology",
            "ac capacity", "dc capacity", "mwac", "mwdc", "inverter manufacturer",
            "module manufacturer"
        },
        "entities": ["plant", "site", "project"],
        "attributes": ["name", "location", "capacity", "commissioning_date", "owner"],
        "domains": ["solar", "bess", "wind", "grid"],
        "content_types": ["metadata"],
    },
    "equipment_schedule": {
        "sheet_keywords": {
            "equipment", "schedule", "device", "asset", "bom", "inventory"
        },
        "header_keywords": {
            "tag", "equipment", "device", "manufacturer", "vendor", "model",
            "serial", "qty", "quantity", "count", "rating", "capacity",
            "description", "type"
        },
        "entities": ["equipment"],
        "attributes": ["tag", "manufacturer", "model", "quantity", "rating"],
        "domains": ["solar", "electrical"],
        "content_types": ["table", "schedule"],
    },
    "inverter_list": {
        "sheet_keywords": {
            "inverter", "inverters", "pcs"
        },
        "header_keywords": {
            "inverter", "inv", "pcs", "skid", "manufacturer", "vendor",
            "model", "serial", "ac power", "dc power", "kwac", "kwdc",
            "station", "block", "quantity"
        },
        "entities": ["inverter"],
        "attributes": ["manufacturer", "model", "ac_power", "dc_power", "quantity"],
        "domains": ["solar", "bess"],
        "content_types": ["table", "equipment_list"],
    },
    "tracker_list": {
        "sheet_keywords": {
            "tracker", "trackers"
        },
        "header_keywords": {
            "tracker", "row", "table", "motor", "controller", "manufacturer",
            "model", "quantity", "axis"
        },
        "entities": ["tracker"],
        "attributes": ["manufacturer", "model", "quantity", "axis"],
        "domains": ["solar"],
        "content_types": ["table", "equipment_list"],
    },
    "string_table": {
        "sheet_keywords": {
            "string", "strings", "combiner", "dc"
        },
        "header_keywords": {
            "string", "combiner", "mppt", "dc", "module", "modules per string",
            "strings per inverter", "input", "channel", "home run", "array"
        },
        "entities": ["string", "combiner", "array"],
        "attributes": ["string_count", "modules_per_string", "mppt", "dc_input"],
        "domains": ["solar"],
        "content_types": ["table", "electrical"],
    },
}


def summarize_xlsx(prof: DocProfile) -> FileSummary:
    sections: list[FileSection] = []
    domains = set()
    content_types = set()
    entities = set()
    attributes = set()
    relationships = set()

    try:
        wb = load_workbook(prof.path, read_only=True, data_only=True)
    except Exception:
        return FileSummary(
            file_name=prof.name,
            file_type="xlsx",
            path=prof.path,
            sections=[],
            domains=[],
            content_types=[],
            entities=[],
            attributes=[],
            relationships=[],
        )

    for ws in wb.worksheets:
        sheet_summary = classify_sheet(ws)
        sections.append(
            FileSection(
                label=sheet_summary["label"],
                sheet_name=ws.title,
                confidence=sheet_summary["confidence"],
                signals=sheet_summary["signals"],
                domains=sheet_summary["domains"],
                content_types=sheet_summary["content_types"],
                entities=sheet_summary["entities"],
                attributes=sheet_summary["attributes"],
                relationships=[],
            )
        )

        domains.update(sheet_summary["domains"])
        content_types.update(sheet_summary["content_types"])
        entities.update(sheet_summary["entities"])
        attributes.update(sheet_summary["attributes"])

    return FileSummary(
        file_name=prof.name,
        file_type="xlsx",
        path=prof.path,
        sections=sections,
        domains=sorted(domains),
        content_types=sorted(content_types),
        entities=sorted(entities),
        attributes=sorted(attributes),
        relationships=sorted(relationships),
    )


def classify_sheet(ws) -> dict:
    sheet_name = normalize_text(ws.title)
    headers = extract_header_candidates(ws)
    header_text = " ".join(headers)

    scores: dict[str, float] = {}
    matched_signals: dict[str, list[str]] = {}

    for label, rule in CATEGORY_RULES.items():
        score = 0.0
        signals: list[str] = []

        for kw in rule["sheet_keywords"]:
            if kw in sheet_name:
                score += 2.5
                signals.append(f"sheet_name:{kw}")

        for kw in rule["header_keywords"]:
            if kw in header_text:
                score += 1.0
                signals.append(f"header:{kw}")

        structure_score = score_structure(label, headers)
        if structure_score > 0:
            score += structure_score
            signals.append(f"structure:{structure_score:.1f}")

        scores[label] = score
        matched_signals[label] = signals

    best_label, best_score = max(scores.items(), key=lambda x: x[1])

    if best_score < 2.0:
        return {
            "label": "unknown",
            "confidence": 0.25,
            "signals": [],
            "domains": [],
            "content_types": [],
            "entities": [],
            "attributes": [],
        }

    confidence = compute_confidence(best_score, scores)
    rule = CATEGORY_RULES[best_label]

    return {
        "label": best_label,
        "confidence": confidence,
        "signals": matched_signals[best_label][:8],
        "domains": rule["domains"],
        "content_types": rule["content_types"],
        "entities": rule["entities"],
        "attributes": rule["attributes"],
    }


def extract_header_candidates(ws, max_rows: int = 8, max_cols: int = 20) -> list[str]:
    """
    Read the top few rows and pick the row that looks most like a header row.
    """
    rows: list[list[str]] = []

    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
        cleaned = [normalize_text(v) for v in row if normalize_text(v)]
        if cleaned:
            rows.append(cleaned)

    if not rows:
        return []

    scored_rows = []
    for row in rows:
        unique_count = len(set(row))
        alpha_like = sum(1 for c in row if has_letters(c))
        scored_rows.append((unique_count + alpha_like, row))

    scored_rows.sort(key=lambda x: x[0], reverse=True)
    return scored_rows[0][1]


def score_structure(label: str, headers: list[str]) -> float:
    if not headers:
        return 0.0

    header_set = set(headers)
    width = len(headers)

    if label == "plant_metadata":
        if width <= 6:
            return 0.8
        if {"field", "value"} & header_set:
            return 1.0

    if label in {"equipment_schedule", "inverter_list", "tracker_list", "string_table"}:
        if width >= 4:
            return 0.8
        if {"manufacturer", "model"} <= header_set:
            return 1.0
        if {"qty", "quantity"} & header_set:
            return 0.5

    return 0.0


def compute_confidence(best_score: float, scores: dict[str, float]) -> float:
    ordered = sorted(scores.values(), reverse=True)
    second = ordered[1] if len(ordered) > 1 else 0.0
    margin = best_score - second

    if best_score >= 8 and margin >= 3:
        return 0.95
    if best_score >= 6 and margin >= 2:
        return 0.88
    if best_score >= 4:
        return 0.78
    return 0.62


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    text = " ".join(text.split())
    return text


def has_letters(text: str) -> bool:
    return any(ch.isalpha() for ch in text)